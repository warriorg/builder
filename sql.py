#coding=utf-8
import sys
import os
from openpyxl import load_workbook
from db.column import Column

#读取excel的数据
def read_excel(excel_name):
    #打开一个workbook
    wb = load_workbook(filename=excel_name)
    #获取所有表格（worksheet）的名字
    sheets = wb.sheetnames
    
    os.remove(excel_name + '.sql')

    #遍历每一个sheet，并且拿到worksheet对象
    for i in range(len(sheets)):
        sheet = wb[sheets[i]]

        if sheet['b1'].value == None:
            continue

        table_name = sheet['b1'].value
        table_name_comment = sheet['e1'].value

        table_sql = 'CREATE TABLE "' + table_name + '" ('
        common_sql = ''
        for rowNum in range(5, sheet.max_row + 1):
            if (sheet.cell(row=rowNum, column=1).value != None):
                table_sql += '\n\t'
                common_sql += '\nCOMMENT ON COLUMN "' + table_name + '".'
            if (sheet.cell(row=rowNum, column=2).value != None):
                table_sql += '"' + sheet.cell(row=rowNum, column=2).value + '" '
                common_sql += '"' + sheet.cell(row=rowNum, column=2).value + '" IS '
            if (sheet.cell(row=rowNum, column=3).value != None):
                db_type = sheet.cell(row=rowNum, column=3).value 
                if db_type == 'VARCHAR':
                    db_type = 'VARCHAR2' 
                table_sql += db_type
            if (sheet.cell(row=rowNum, column=4).value != None):
                db_type = sheet.cell(row=rowNum, column=3).value 
                if db_type == 'DATE':
                    table_sql += ','
                else:
                    table_sql += '(' + str(sheet.cell(row=rowNum, column=4).value) + '),'
            if (sheet.cell(row=rowNum, column=1).value != None):
                common_sql += "'" + sheet.cell(row=rowNum, column=1).value + "';" 
            
        table_sql += '\n\tPRIMARY KEY ("SID")\n);'
        common_sql += '\nCOMMENT ON TABLE "' + table_name + '" IS \'' + table_name_comment + '\';\n\n'

        with open(excel_name + '.sql', 'a') as outfile: 
            outfile.write(table_sql)
            outfile.write(common_sql)
        
 
#将读取的数据存入txt文档中
def text_save(content,filename,mode='a'):
    #打开文件
    file = open(filename,mode)
    for i in range(len(content)):
        number = content[i] + '\n'
        file.write(number)
    file.close()
 
if __name__ == '__main__':
    args = sys.argv[1:]
    excel_name = args[0].upper()
    read_excel(excel_name)
